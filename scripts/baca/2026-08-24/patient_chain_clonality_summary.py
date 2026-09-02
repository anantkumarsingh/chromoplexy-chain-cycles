"""
One row per patient, pulling together:
  - ETS / CHD1 status (mmc5 Table S5B)
  - Total breakpoints + DSB distribution per chromosome, computed from
    the whole-genome rearrangement data (chrom_aberrations_baca.csv,
    == mmc3 Table S3C)
  - 'Maximum chromosomes in one chain', copied as-is from mmc5 Table S5B
  - Deletion bridge partners across ALL of a patient's mmc5 chains
    (Table S5A), listed per chain with breakpoint number + strand
  - Clonality info from mmc6 Table S6B, as-is (not interpreted)

Scope: the 57 patients in clinical_phenotypes.csv. Table S5B also
contains other-cohort (BR-/LUAD-/ME-/HN_ prefixed) and simulated/
scrambled rows sharing the same 'Individual' column — filtered out by
restricting to known_patients, same pattern as
core/baca_aberration_clinical_analysis.py's load_ets_status_from_mmc5().

Output:
  - results/2026-08-24/patient_chain_clonality_summary.xlsx (formatted:
    bold header row, bold chromosome labels inside the breaks-per-
    chromosome cell, real multi-line cells for deletion bridges and
    clonality — the "pretty" deliverable)
  - results/2026-08-24/patient_chain_clonality_summary.csv (same data,
    plain text — semicolon-spaced breaks-per-chromosome, newline-
    separated deletion-bridge/clonality entries within each cell)
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter

BACA_DATASET_FOLDER = "/Users/anantkumarsingh/projects/prostate_cancer/nih-tcga-prad/data/baca_dataset"
RESULTS_DIR = "/Users/anantkumarsingh/projects/prostate_cancer/nih-tcga-prad/results/2026-08-24"

CLINICAL_PATH = f"{BACA_DATASET_FOLDER}/clinical_phenotypes.csv"
CHROM_PATH = f"{BACA_DATASET_FOLDER}/chrom_aberrations_baca.csv"
MMC5_PATH = f"{BACA_DATASET_FOLDER}/mmc5.xlsx"
MMC6_PATH = f"{BACA_DATASET_FOLDER}/mmc6.xlsx"

STRAND_MAP = {'Forward': '+', 'Reverse': '-'}
ETS_MAP = {'POSITIVE': 'ETS+', 'NEGATIVE': 'ETS-'}
CHD1_MAP = {'WT': 'CHD1wt', '-': 'CHD1del'}

COLUMNS = [
    'Individual', 'ETS_status', 'CHD1_status', 'Total_breakpoints',
    'Breaks_per_chromosome', 'Maximum chromosomes in one chain',
    'Deletion_bridges_present', 'Deletion_bridges_detail',
    'clonality_status_per_gene',
]


def load_patients():
    clinical = pd.read_csv(CLINICAL_PATH)
    return sorted(clinical['Individual'].unique().tolist())


def load_s5b(patients):
    s5b = pd.read_excel(MMC5_PATH, sheet_name='Table S5B')
    s5b = s5b[s5b['Individual'].isin(set(patients))].copy()
    return s5b.set_index('Individual')


def load_s5a(patients):
    s5a = pd.read_excel(MMC5_PATH, sheet_name='Table S5A')
    s5a.columns = [c.strip() for c in s5a.columns]
    s5a = s5a[s5a['Individual'].isin(set(patients))].copy()
    s5a['Strand'] = s5a['Strand'].map(STRAND_MAP)
    s5a['Deletion bridge partner breakpoint'] = pd.to_numeric(
        s5a['Deletion bridge partner breakpoint'], errors='coerce'
    )
    return s5a


def compute_breaks_per_chromosome(chrom_df, patient):
    """Returns (total_breakpoints, [(chrom_int, count), ...] sorted ascending)."""
    p = chrom_df[chrom_df['Individual'] == patient]
    combined = pd.concat([
        p['Breakpoint 1 chromosome'].dropna(),
        p['Breakpoint 2 chromosome'].dropna(),
    ]).astype(int)
    counts = combined.value_counts().sort_index()
    return len(combined), list(counts.items())


def build_deletion_bridge_detail(patient_s5a):
    """Returns (Yes/No, [ 'Chain N: pair; pair; ...' , ... ] one string per chain)."""
    db_rows = patient_s5a[patient_s5a['Deletion bridge partner breakpoint'].notna()]
    if db_rows.empty:
        return "No", []

    bp_lookup = patient_s5a.set_index('Breakpoint number')[['Strand', 'Chain number']]

    seen_pairs = set()
    chain_pairs = {}  # chain_number -> list of (bp_lo, strand_lo, bp_hi, strand_hi)

    for _, row in db_rows.iterrows():
        bp_a = int(row['Breakpoint number'])
        bp_b = int(row['Deletion bridge partner breakpoint'])
        pair_key = tuple(sorted([bp_a, bp_b]))
        if pair_key in seen_pairs:
            continue
        seen_pairs.add(pair_key)

        chain_num = int(row['Chain number'])
        strand_a = row['Strand']
        strand_b = bp_lookup.loc[bp_b, 'Strand'] if bp_b in bp_lookup.index else '?'

        lo, hi = pair_key
        strand_lo, strand_hi = (strand_a, strand_b) if bp_a == lo else (strand_b, strand_a)

        chain_pairs.setdefault(chain_num, []).append((lo, strand_lo, hi, strand_hi))

    chain_lines = []
    for chain_num in sorted(chain_pairs.keys()):
        pairs = sorted(chain_pairs[chain_num])
        pair_strs = [f"BP{a}({sa}) --- BP{b}({sb})" for a, sa, b, sb in pairs]
        chain_lines.append(f"Chain {chain_num}: " + "; ".join(pair_strs))

    return "Yes", chain_lines


def build_clonality_list(s6b_df, patient):
    """Returns a list of 'GENE:CN=x,Clon=y%[range]' strings, one per gene."""
    rows = s6b_df[s6b_df['Sample'] == patient]
    if rows.empty:
        return []
    out = []
    for _, r in rows.sort_values('Gene').iterrows():
        out.append(
            f"{r['Gene']}:CN={r['Somatic Copy Number']},"
            f"Clon={r['Clonality %']}%{r['Clonality % Ranges']}"
        )
    return out


def collect_patient_rows():
    patients = load_patients()
    s5b = load_s5b(patients)
    s5a = load_s5a(patients)
    chrom_df = pd.read_csv(CHROM_PATH)
    s6b = pd.read_excel(MMC6_PATH, sheet_name='Table S6B')

    records = []
    for patient in patients:
        s5b_row = s5b.loc[patient] if patient in s5b.index else None

        total_bp, chrom_counts = compute_breaks_per_chromosome(chrom_df, patient)
        patient_s5a = s5a[s5a['Individual'] == patient]
        db_yn, db_lines = build_deletion_bridge_detail(patient_s5a)
        clonality_lines = build_clonality_list(s6b, patient)

        records.append({
            'Individual': patient,
            'ETS_status': ETS_MAP.get(s5b_row['ETS fusion']) if s5b_row is not None else None,
            'CHD1_status': CHD1_MAP.get(s5b_row['CHD1']) if s5b_row is not None else None,
            'Total_breakpoints': total_bp,
            'chrom_counts': chrom_counts,  # [(chrom_int, count), ...]
            'Maximum chromosomes in one chain': s5b_row['Maximum chromosomes in one chain'] if s5b_row is not None else None,
            'Deletion_bridges_present': db_yn,
            'db_lines': db_lines,          # list of strings, one per chain
            'clonality_lines': clonality_lines,  # list of strings, one per gene
        })
    return records


def write_csv(records, out_path):
    rows = []
    for r in records:
        breaks_str = "; ".join(f"CHR{c} - {n}" for c, n in r['chrom_counts'])
        db_str = "\n".join(r['db_lines'])
        clon_str = "\n".join(r['clonality_lines'])
        rows.append({
            'Individual': r['Individual'],
            'ETS_status': r['ETS_status'],
            'CHD1_status': r['CHD1_status'],
            'Total_breakpoints': r['Total_breakpoints'],
            'Breaks_per_chromosome': breaks_str,
            'Maximum chromosomes in one chain': r['Maximum chromosomes in one chain'],
            'Deletion_bridges_present': r['Deletion_bridges_present'],
            'Deletion_bridges_detail': db_str,
            'clonality_status_per_gene': clon_str,
        })
    pd.DataFrame(rows, columns=COLUMNS).to_csv(out_path, index=False)


def write_xlsx(records, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Summary"

    header_font = Font(bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top", horizontal="left")
    bold_inline = InlineFont(b=True)

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="left")

    for r in records:
        row_idx = ws.max_row + 1

        # Rich text for Breaks_per_chromosome: bold "CHRn", regular " - count; "
        rich_parts = []
        for i, (c, n) in enumerate(r['chrom_counts']):
            rich_parts.append(TextBlock(bold_inline, f"CHR{c}"))
            sep = "; " if i < len(r['chrom_counts']) - 1 else ""
            rich_parts.append(f" - {n}{sep}")
        breaks_rich = CellRichText(rich_parts) if rich_parts else ""

        db_str = "\n".join(r['db_lines'])
        clon_str = "\n".join(r['clonality_lines'])

        row_values = [
            r['Individual'],
            r['ETS_status'],
            r['CHD1_status'],
            r['Total_breakpoints'],
            breaks_rich,
            r['Maximum chromosomes in one chain'],
            r['Deletion_bridges_present'],
            db_str,
            clon_str,
        ]
        ws.append(row_values)

        for cell in ws[row_idx]:
            cell.alignment = wrap_top

        n_lines = max(1, len(r['db_lines']), len(r['clonality_lines']), 1)
        ws.row_dimensions[row_idx].height = max(15, n_lines * 15)

    widths = {
        'A': 14, 'B': 12, 'C': 13, 'D': 16, 'E': 55,
        'F': 14, 'G': 14, 'H': 55, 'I': 45,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    wb.save(out_path)


def main():
    os.makedirs(RESULTS_DIR, exist_ok=True)
    records = collect_patient_rows()

    csv_path = f"{RESULTS_DIR}/patient_chain_clonality_summary.csv"
    xlsx_path = f"{RESULTS_DIR}/patient_chain_clonality_summary.xlsx"

    write_csv(records, csv_path)
    write_xlsx(records, xlsx_path)

    print(f"Wrote {csv_path} — {len(records)} patients")
    print(f"Wrote {xlsx_path} — {len(records)} patients")


if __name__ == '__main__':
    main()
